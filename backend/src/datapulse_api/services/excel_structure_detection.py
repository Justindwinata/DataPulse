from io import BytesIO
from typing import Any
from collections import Counter

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency is declared, fallback is defensive.
    xlrd = None

from datapulse_api.models import (
    ExcelSheetMetadata,
    ExcelWorkbookMetadata,
    RawTablePreview,
    StructureDetectionResult,
    StructureDetectionStatus,
    StructureNextStep,
    StructureWarning,
    StructureWarningSeverity,
)
from datapulse_api.services.file_validation import detect_extension, sanitize_filename

EXCEL_EXTENSIONS = frozenset({"xlsx", "xls"})
MAX_EXCEL_SAMPLE_ROWS = 50
MAX_EXCEL_PREVIEW_ROWS = 20
EXCEL_LIMITATION_WARNING = StructureWarning(
    code="excel_formatting_not_preserved",
    message=(
        "Excel formatting, formulas, merged cell behavior, charts, and pivot tables "
        "are not preserved in preview."
    ),
    severity=StructureWarningSeverity.INFO,
)


def detect_excel_workbook(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
) -> StructureDetectionResult:
    safe_filename = sanitize_filename(filename)
    extension = detect_extension(safe_filename) or "unknown"

    if extension not in EXCEL_EXTENSIONS:
        return _excel_rejected_result(
            filename=filename,
            safe_filename=safe_filename,
            extension=extension,
            content_type=content_type,
            file_size_bytes=len(content),
            code="file_not_excel",
            message="Excel sheet discovery supports XLSX and XLS files.",
            status=StructureDetectionStatus.REJECTED,
        )

    if not content:
        return _excel_rejected_result(
            filename=filename,
            safe_filename=safe_filename,
            extension=extension,
            content_type=content_type,
            file_size_bytes=0,
            code="empty_file",
            message="File is empty. Upload a non-empty Excel workbook.",
            status=StructureDetectionStatus.REJECTED,
        )

    try:
        workbook_metadata = (
            _discover_xlsx_workbook(content)
            if extension == "xlsx"
            else _discover_xls_workbook(content)
        )
    except Exception as error:
        return _excel_rejected_result(
            filename=filename,
            safe_filename=safe_filename,
            extension=extension,
            content_type=content_type,
            file_size_bytes=len(content),
            code="excel_workbook_unreadable",
            message="Excel workbook could not be read as a supported table-like workbook.",
            status=StructureDetectionStatus.REJECTED,
        )

    return StructureDetectionResult(
        original_filename=filename,
        safe_filename=safe_filename,
        detected_extension=extension,
        file_size_bytes=len(content),
        content_type=content_type,
        structure_status=StructureDetectionStatus.SHEET_SELECTION_REQUIRED,
        workbook=workbook_metadata,
        warnings=[EXCEL_LIMITATION_WARNING],
        next_step=StructureNextStep.SELECT_EXCEL_SHEET,
    )


def detect_excel_sheet_preview(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    sheet_name: str,
) -> StructureDetectionResult:
    safe_filename = sanitize_filename(filename)
    extension = detect_extension(safe_filename) or "unknown"

    if extension not in EXCEL_EXTENSIONS:
        return _excel_rejected_result(
            filename=filename,
            safe_filename=safe_filename,
            extension=extension,
            content_type=content_type,
            file_size_bytes=len(content),
            code="file_not_excel",
            message="Excel sheet preview supports XLSX and XLS files.",
            status=StructureDetectionStatus.REJECTED,
        )
    if not content:
        return _excel_rejected_result(
            filename=filename,
            safe_filename=safe_filename,
            extension=extension,
            content_type=content_type,
            file_size_bytes=0,
            code="empty_file",
            message="File is empty. Upload a non-empty Excel workbook.",
            status=StructureDetectionStatus.REJECTED,
        )

    try:
        workbook_metadata = (
            _discover_xlsx_workbook(content)
            if extension == "xlsx"
            else _discover_xls_workbook(content)
        )
        if sheet_name not in workbook_metadata.sheet_names:
            return _excel_rejected_result(
                filename=filename,
                safe_filename=safe_filename,
                extension=extension,
                content_type=content_type,
                file_size_bytes=len(content),
                code="selected_sheet_not_found",
                message="Selected sheet was not found in the uploaded workbook.",
                status=StructureDetectionStatus.REJECTED,
            )
        rows = (
            _sample_xlsx_sheet(content, sheet_name)
            if extension == "xlsx"
            else _sample_xls_sheet(content, sheet_name)
        )
    except Exception:
        return _excel_rejected_result(
            filename=filename,
            safe_filename=safe_filename,
            extension=extension,
            content_type=content_type,
            file_size_bytes=len(content),
            code="excel_workbook_unreadable",
            message="Excel workbook could not be read as a supported table-like workbook.",
            status=StructureDetectionStatus.REJECTED,
        )

    warnings = [EXCEL_LIMITATION_WARNING]
    warnings.extend(_detect_excel_sample_warnings(rows))
    meaningful_rows = [(index, row) for index, row in enumerate(rows) if _row_has_value(row)]

    if not meaningful_rows:
        warnings.append(
            StructureWarning(
                code="empty_sheet",
                message="Selected sheet is empty in the sampled range.",
                severity=StructureWarningSeverity.ERROR,
            )
        )
        return StructureDetectionResult(
            original_filename=filename,
            safe_filename=safe_filename,
            detected_extension=extension,
            file_size_bytes=len(content),
            content_type=content_type,
            structure_status=StructureDetectionStatus.REJECTED,
            workbook=workbook_metadata,
            selected_sheet_name=sheet_name,
            warnings=_deduplicate_warnings(warnings),
            next_step=StructureNextStep.SELECT_EXCEL_SHEET,
        )

    header_index, has_header = _detect_header_row(meaningful_rows)
    sheet_rows = [(index, row) for index, row in meaningful_rows if index >= (header_index or 0)]
    header_row = next((row for index, row in sheet_rows if index == header_index), None)
    column_count = len(header_row) if has_header and header_row is not None else _detect_column_count(sheet_rows)
    column_names, generated_column_names, name_warnings = _build_column_names(
        header_row=header_row if has_header else None,
        column_count=column_count,
    )
    warnings.extend(name_warnings)
    warnings.extend(_detect_empty_column_warnings(sheet_rows, column_count))

    data_rows = [(index, row) for index, row in sheet_rows if not has_header or index != header_index]
    preview_rows = [_normalize_row(row, column_count) for _, row in data_rows[:MAX_EXCEL_PREVIEW_ROWS]]

    if len(meaningful_rows) < 2:
        warnings.append(
            StructureWarning(
                code="very_small_sheet",
                message="The selected sheet sample contains fewer than two meaningful rows.",
                severity=StructureWarningSeverity.INFO,
            )
        )

    return StructureDetectionResult(
        original_filename=filename,
        safe_filename=safe_filename,
        detected_extension=extension,
        file_size_bytes=len(content),
        content_type=content_type,
        structure_status=StructureDetectionStatus.DETECTED,
        workbook=workbook_metadata,
        selected_sheet_name=sheet_name,
        has_detected_header=has_header,
        header_row_index=header_index,
        column_names=column_names,
        generated_column_names=generated_column_names,
        detected_column_count=column_count,
        sampled_row_count=len(rows),
        preview_row_count=len(preview_rows),
        total_row_count_calculated=False,
        warnings=_deduplicate_warnings(warnings),
        preview=RawTablePreview(columns=column_names, rows=preview_rows),
        next_step=StructureNextStep.QUALITY_ISSUE_DETECTION,
    )


def _discover_xlsx_workbook(content: bytes) -> ExcelWorkbookMetadata:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    sheets = [
        ExcelSheetMetadata(
            sheet_name=worksheet.title,
            sheet_index=index,
            max_row=worksheet.max_row or 0,
            max_column=worksheet.max_column or 0,
            is_empty=_xlsx_sheet_is_empty(worksheet),
        )
        for index, worksheet in enumerate(workbook.worksheets)
    ]
    return ExcelWorkbookMetadata(
        workbook_type="xlsx",
        sheet_names=sheet_names,
        sheet_count=len(sheet_names),
        default_sheet_name=sheet_names[0],
        sheets=sheets,
    )


def _discover_xls_workbook(content: bytes) -> ExcelWorkbookMetadata:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed")
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    sheet_names = workbook.sheet_names()
    sheets = [
        ExcelSheetMetadata(
            sheet_name=sheet_name,
            sheet_index=index,
            max_row=workbook.sheet_by_index(index).nrows,
            max_column=workbook.sheet_by_index(index).ncols,
            is_empty=workbook.sheet_by_index(index).nrows == 0
            or workbook.sheet_by_index(index).ncols == 0,
        )
        for index, sheet_name in enumerate(sheet_names)
    ]
    return ExcelWorkbookMetadata(
        workbook_type="xls",
        sheet_names=sheet_names,
        sheet_count=len(sheet_names),
        default_sheet_name=sheet_names[0],
        sheets=sheets,
    )


def _sample_xlsx_sheet(content: bytes, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(max_row=MAX_EXCEL_SAMPLE_ROWS, values_only=True):
        rows.append([_cell_to_text(cell) for cell in row])
    return rows


def _sample_xls_sheet(content: bytes, sheet_name: str) -> list[list[str]]:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed")
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    sheet = workbook.sheet_by_name(sheet_name)
    rows: list[list[str]] = []
    for row_index in range(min(sheet.nrows, MAX_EXCEL_SAMPLE_ROWS)):
        rows.append([_cell_to_text(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)])
    return rows


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_has_value(row: list[str]) -> bool:
    return any(cell.strip() for cell in row)


def _detect_excel_sample_warnings(rows: list[list[str]]) -> list[StructureWarning]:
    warnings: list[StructureWarning] = []
    if len(rows) >= MAX_EXCEL_SAMPLE_ROWS:
        warnings.append(
            StructureWarning(
                code="sample_row_limit_reached",
                message="Only the first bounded sample of Excel rows was inspected.",
                severity=StructureWarningSeverity.INFO,
            )
        )
    if any(not _row_has_value(row) for row in rows):
        warnings.append(
            StructureWarning(
                code="empty_rows_in_sample",
                message="Empty rows were detected in the selected sheet sample.",
            )
        )
    return warnings


def _detect_header_row(rows: list[tuple[int, list[str]]]) -> tuple[int | None, bool]:
    if not rows:
        return None, False
    for row_index, row in rows[:5]:
        non_empty = [cell for cell in row if cell.strip()]
        if len(non_empty) >= 2 and _looks_like_column_labels(non_empty):
            return row_index, True
    return rows[0][0], False


def _looks_like_column_labels(cells: list[str]) -> bool:
    text_like_count = sum(
        bool(cell.strip()) and not _is_numeric(cell) and not any(character.isdigit() for character in cell)
        for cell in cells
    )
    return text_like_count / len(cells) >= 0.8


def _is_numeric(value: str) -> bool:
    normalized = value.strip().replace(",", "")
    if not normalized:
        return False
    try:
        float(normalized)
    except ValueError:
        return False
    return True


def _detect_column_count(rows: list[tuple[int, list[str]]]) -> int:
    widths = [len(_trim_trailing_empty_cells(row)) for _, row in rows if _row_has_value(row)]
    if not widths:
        return 1
    return max(Counter(widths).most_common(1)[0][0], 1)


def _build_column_names(
    *,
    header_row: list[str] | None,
    column_count: int,
) -> tuple[list[str], bool, list[StructureWarning]]:
    warnings: list[StructureWarning] = []
    if header_row is None:
        return [f"column_{index}" for index in range(1, column_count + 1)], True, warnings

    normalized_header = _normalize_row(header_row, column_count)
    names = [
        _normalize_column_name(name) if name.strip() else f"column_{index + 1}"
        for index, name in enumerate(normalized_header)
    ]
    generated = any(not name.strip() for name in normalized_header)
    if generated:
        warnings.append(
            StructureWarning(
                code="missing_column_names",
                message="Missing column names were replaced with generated names.",
            )
        )
    if any(count > 1 for count in Counter(names).values()):
        warnings.append(
            StructureWarning(
                code="duplicate_column_names",
                message="Duplicate column names were detected in the header row.",
            )
        )
    return names, generated, warnings


def _normalize_column_name(name: str) -> str:
    normalized = "_".join(name.strip().lower().split())
    return normalized or "column"


def _detect_empty_column_warnings(
    rows: list[tuple[int, list[str]]],
    column_count: int,
) -> list[StructureWarning]:
    if not rows:
        return []
    empty_columns = []
    for column_index in range(column_count):
        if all(
            column_index >= len(row) or not row[column_index].strip()
            for _, row in rows
        ):
            empty_columns.append(column_index)
    if empty_columns:
        return [
            StructureWarning(
                code="empty_columns_in_sample",
                message="Fully empty columns were detected in the selected sheet sample.",
            )
        ]
    return []


def _trim_trailing_empty_cells(row: list[str]) -> list[str]:
    trimmed = list(row)
    while trimmed and not trimmed[-1].strip():
        trimmed.pop()
    return trimmed or [""]


def _normalize_row(row: list[str], column_count: int) -> list[str]:
    normalized = row[:column_count]
    if len(normalized) < column_count:
        normalized.extend([""] * (column_count - len(normalized)))
    return normalized


def _deduplicate_warnings(warnings: list[StructureWarning]) -> list[StructureWarning]:
    seen: set[str] = set()
    unique: list[StructureWarning] = []
    for warning in warnings:
        if warning.code in seen:
            continue
        seen.add(warning.code)
        unique.append(warning)
    return unique


def _xlsx_sheet_is_empty(worksheet: Any) -> bool:
    for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
        return not any(cell is not None and str(cell).strip() for cell in row)
    return True


def _excel_rejected_result(
    *,
    filename: str,
    safe_filename: str,
    extension: str,
    content_type: str | None,
    file_size_bytes: int,
    code: str,
    message: str,
    status: StructureDetectionStatus,
) -> StructureDetectionResult:
    return StructureDetectionResult(
        original_filename=filename,
        safe_filename=safe_filename,
        detected_extension=extension,
        file_size_bytes=file_size_bytes,
        content_type=content_type,
        structure_status=status,
        warnings=[
            StructureWarning(
                code=code,
                message=message,
                severity=StructureWarningSeverity.ERROR,
            )
        ],
        next_step=StructureNextStep.UPLOAD_SUPPORTED_FILE,
    )
