import csv
import json
import re
from collections import Counter
from io import BytesIO, StringIO
from typing import Iterable

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None

from datapulse_api.models import (
    CleanedTablePreview,
    CleaningNextStep,
    CleaningPreviewResult,
    CleaningPreviewWarning,
    CleaningRuleCode,
    CleaningRuleEffect,
    CleaningRuleEffectStatus,
    CleaningStatus,
    CleaningSummaryAfter,
    CleaningSummaryBefore,
    CleaningWarningSeverity,
    StructureDetectionResult,
    StructureDetectionStatus,
)
from datapulse_api.services.csv_structure_detection import (
    CSV_LIKE_EXTENSIONS,
    EXCEL_EXTENSIONS,
    detect_csv_like_structure,
)
from datapulse_api.services.excel_structure_detection import (
    detect_excel_sheet_preview,
    detect_excel_workbook,
)
from datapulse_api.services.file_validation import detect_extension, sanitize_filename

MAX_CLEANING_SAMPLE_ROWS = 50
MAX_CLEANING_PREVIEW_ROWS = 20
GENERIC_COLUMN_RE = re.compile(r"^column_\d+$")
RULE_LABELS = {
    CleaningRuleCode.TRIM_WHITESPACE: "Trim whitespace",
    CleaningRuleCode.REMOVE_EMPTY_ROWS: "Remove empty rows",
    CleaningRuleCode.REMOVE_DUPLICATE_ROWS: "Remove duplicate rows",
    CleaningRuleCode.DROP_EMPTY_COLUMNS: "Drop empty columns",
    CleaningRuleCode.STANDARDIZE_COLUMN_NAMES: "Standardize column names",
    CleaningRuleCode.GENERATE_MISSING_COLUMN_NAMES: "Generate missing column names",
}


def parse_cleaning_rules(raw_rules: str | list[str] | None) -> list[CleaningRuleCode]:
    if raw_rules is None:
        return []
    values: list[str] = []
    if isinstance(raw_rules, str):
        try:
            decoded = json.loads(raw_rules)
            values = decoded if isinstance(decoded, list) else [str(decoded)]
        except json.JSONDecodeError:
            values = [raw_rules]
    else:
        for item in raw_rules:
            try:
                decoded = json.loads(item)
                if isinstance(decoded, list):
                    values.extend(str(value) for value in decoded)
                else:
                    values.append(str(decoded))
            except json.JSONDecodeError:
                values.append(item)

    rules: list[CleaningRuleCode] = []
    for value in values:
        rule = CleaningRuleCode(value)
        if rule not in rules:
            rules.append(rule)
    return rules


def generate_cleaning_preview(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    rules: Iterable[CleaningRuleCode],
    sheet_name: str | None = None,
) -> CleaningPreviewResult:
    safe_filename = sanitize_filename(filename)
    extension = detect_extension(safe_filename) or "unknown"
    selected_rules = list(rules)

    if extension in CSV_LIKE_EXTENSIONS:
        structure = detect_csv_like_structure(
            filename=filename,
            content_type=content_type,
            content=content,
        )
        rows = _extract_csv_rows(content=content, structure=structure)
        return _preview_from_structure(structure, rows, selected_rules)

    if extension in EXCEL_EXTENSIONS:
        if not sheet_name:
            workbook_result = detect_excel_workbook(
                filename=filename,
                content_type=content_type,
                content=content,
            )
            if workbook_result.structure_status == StructureDetectionStatus.SHEET_SELECTION_REQUIRED:
                return _sheet_selection_required_result(workbook_result)
            return _rejected_result_from_structure(workbook_result)
        structure = detect_excel_sheet_preview(
            filename=filename,
            content_type=content_type,
            content=content,
            sheet_name=sheet_name,
        )
        rows = _extract_excel_rows(
            content=content,
            extension=extension,
            sheet_name=sheet_name,
            structure=structure,
        )
        return _preview_from_structure(structure, rows, selected_rules)

    structure = detect_csv_like_structure(
        filename=filename,
        content_type=content_type,
        content=content,
    )
    return _rejected_result_from_structure(structure)


def _preview_from_structure(
    structure: StructureDetectionResult,
    rows: list[list[str]] | None,
    rules: list[CleaningRuleCode],
) -> CleaningPreviewResult:
    if structure.structure_status != StructureDetectionStatus.DETECTED or structure.preview is None:
        return _rejected_result_from_structure(structure)

    columns = list(structure.preview.columns)
    source_rows = rows if rows is not None else structure.preview.rows
    normalized_rows = [_normalize_row(row, len(columns)) for row in source_rows]
    before = _summarize_before(columns, normalized_rows)
    current_columns = list(columns)
    current_rows = [list(row) for row in normalized_rows]
    effects: list[CleaningRuleEffect] = []

    for rule in rules:
        current_columns, current_rows, effect = _apply_rule(rule, current_columns, current_rows)
        effects.append(effect)

    after = CleaningSummaryAfter(
        row_count=len(current_rows),
        column_count=len(current_columns),
        removed_empty_rows_count=before.row_count - len(current_rows)
        if CleaningRuleCode.REMOVE_EMPTY_ROWS in rules
        else 0,
        removed_duplicate_rows_count=_effect_rows(effects, CleaningRuleCode.REMOVE_DUPLICATE_ROWS),
        dropped_empty_columns_count=_effect_columns(effects, CleaningRuleCode.DROP_EMPTY_COLUMNS),
        renamed_columns_count=max(
            _effect_columns(effects, CleaningRuleCode.STANDARDIZE_COLUMN_NAMES),
            _effect_columns(effects, CleaningRuleCode.GENERATE_MISSING_COLUMN_NAMES),
        ),
    )
    warnings = [
        CleaningPreviewWarning(
            code="sample_based_preview",
            message="Cleaning preview is generated from a bounded sample and does not modify your original file.",
        )
    ]
    if not rules:
        warnings.append(
            CleaningPreviewWarning(
                code="no_rules_selected",
                message="No cleaning rules were selected, so the cleaned preview matches the sampled input.",
                severity=CleaningWarningSeverity.WARNING,
            )
        )
    if structure.detected_extension in EXCEL_EXTENSIONS:
        warnings.append(
            CleaningPreviewWarning(
                code="excel_formatting_not_preserved",
                message="Excel formatting, formulas, merged cell behavior, charts, and pivot tables are not preserved in preview.",
            )
        )
    for effect in effects:
        if effect.status == CleaningRuleEffectStatus.NO_EFFECT:
            warnings.append(
                CleaningPreviewWarning(
                    code=f"{effect.rule.value}_no_effect",
                    message=effect.message,
                )
            )

    return CleaningPreviewResult(
        cleaning_status=CleaningStatus.PREVIEW_GENERATED,
        original_filename=structure.original_filename,
        safe_filename=structure.safe_filename,
        detected_extension=structure.detected_extension,
        selected_sheet_name=structure.selected_sheet_name,
        applied_rules=rules,
        before_summary=before,
        after_summary=after,
        rule_effects=effects,
        cleaned_preview=CleanedTablePreview(
            columns=current_columns,
            rows=current_rows[:MAX_CLEANING_PREVIEW_ROWS],
        ),
        warnings=warnings,
        next_step=CleaningNextStep.DOWNLOAD_CLEANED_CSV,
    )


def _apply_rule(
    rule: CleaningRuleCode,
    columns: list[str],
    rows: list[list[str]],
) -> tuple[list[str], list[list[str]], CleaningRuleEffect]:
    if rule == CleaningRuleCode.TRIM_WHITESPACE:
        changed_cells = 0
        affected_rows: set[int] = set()
        affected_columns: set[int] = set()
        cleaned_rows = []
        for row_index, row in enumerate(rows):
            cleaned_row = []
            for column_index, cell in enumerate(row):
                trimmed = cell.strip()
                if trimmed != cell:
                    changed_cells += 1
                    affected_rows.add(row_index)
                    affected_columns.add(column_index)
                cleaned_row.append(trimmed)
            cleaned_rows.append(cleaned_row)
        return columns, cleaned_rows, _effect(
            rule,
            changed_cells > 0,
            f"Trimmed whitespace in {changed_cells} sampled cells."
            if changed_cells
            else "No leading or trailing whitespace was found in the sampled cells.",
            len(affected_rows),
            len(affected_columns),
        )

    if rule == CleaningRuleCode.REMOVE_EMPTY_ROWS:
        cleaned_rows = [row for row in rows if any(cell.strip() for cell in row)]
        removed = len(rows) - len(cleaned_rows)
        return columns, cleaned_rows, _effect(
            rule,
            removed > 0,
            f"Removed {removed} empty rows from the preview sample."
            if removed
            else "No empty rows were found in the preview sample.",
            removed,
            0,
        )

    if rule == CleaningRuleCode.REMOVE_DUPLICATE_ROWS:
        seen: set[tuple[str, ...]] = set()
        cleaned_rows = []
        removed = 0
        for row in rows:
            key = tuple(row)
            if key in seen:
                removed += 1
                continue
            seen.add(key)
            cleaned_rows.append(row)
        return columns, cleaned_rows, _effect(
            rule,
            removed > 0,
            f"Removed {removed} duplicate rows from the preview sample."
            if removed
            else "No exact duplicate rows were found in the preview sample.",
            removed,
            0,
        )

    if rule == CleaningRuleCode.DROP_EMPTY_COLUMNS:
        keep_indexes = [
            index
            for index, _ in enumerate(columns)
            if any(index < len(row) and row[index].strip() for row in rows)
        ]
        dropped = len(columns) - len(keep_indexes)
        cleaned_columns = [columns[index] for index in keep_indexes]
        cleaned_rows = [[row[index] for index in keep_indexes] for row in rows]
        return cleaned_columns, cleaned_rows, _effect(
            rule,
            dropped > 0,
            f"Dropped {dropped} fully empty columns from the preview sample."
            if dropped
            else "No fully empty columns were found in the preview sample.",
            0,
            dropped,
        )

    if rule == CleaningRuleCode.STANDARDIZE_COLUMN_NAMES:
        standardized = [_standardize_column_name(column, index) for index, column in enumerate(columns)]
        unique = _make_unique_columns(standardized)
        renamed = sum(1 for before, after_name in zip(columns, unique, strict=False) if before != after_name)
        return unique, rows, _effect(
            rule,
            renamed > 0,
            f"Standardized {renamed} column names."
            if renamed
            else "Column names already matched the standard style.",
            0,
            renamed,
        )

    if rule == CleaningRuleCode.GENERATE_MISSING_COLUMN_NAMES:
        generated = []
        renamed = 0
        for index, column in enumerate(columns):
            if not column.strip() or GENERIC_COLUMN_RE.match(column):
                generated.append(f"column_{index + 1}")
                renamed += 1
            else:
                generated.append(column)
        generated = _make_unique_columns(generated)
        return generated, rows, _effect(
            rule,
            renamed > 0,
            f"Generated {renamed} missing column names."
            if renamed
            else "No missing column names were found.",
            0,
            renamed,
        )

    return columns, rows, _effect(rule, False, "Rule is not implemented in this milestone.", 0, 0)


def _summarize_before(columns: list[str], rows: list[list[str]]) -> CleaningSummaryBefore:
    return CleaningSummaryBefore(
        row_count=len(rows),
        column_count=len(columns),
        empty_rows_count=sum(1 for row in rows if not any(cell.strip() for cell in row)),
        duplicate_rows_count=_duplicate_row_count(rows),
        empty_columns_count=sum(
            1
            for index, _ in enumerate(columns)
            if not any(index < len(row) and row[index].strip() for row in rows)
        ),
    )


def _duplicate_row_count(rows: list[list[str]]) -> int:
    counts = Counter(tuple(row) for row in rows if any(cell.strip() for cell in row))
    return sum(count - 1 for count in counts.values() if count > 1)


def _effect(
    rule: CleaningRuleCode,
    changed: bool,
    message: str,
    affected_rows: int,
    affected_columns: int,
) -> CleaningRuleEffect:
    return CleaningRuleEffect(
        rule=rule,
        label=RULE_LABELS[rule],
        status=CleaningRuleEffectStatus.APPLIED if changed else CleaningRuleEffectStatus.NO_EFFECT,
        message=message,
        affected_rows=affected_rows,
        affected_columns=affected_columns,
    )


def _effect_rows(effects: list[CleaningRuleEffect], rule: CleaningRuleCode) -> int:
    return next((effect.affected_rows for effect in effects if effect.rule == rule), 0)


def _effect_columns(effects: list[CleaningRuleEffect], rule: CleaningRuleCode) -> int:
    return next((effect.affected_columns for effect in effects if effect.rule == rule), 0)


def _standardize_column_name(column: str, index: int) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", column.strip().lower())
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or f"column_{index + 1}"


def _make_unique_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for column in columns:
        count = seen.get(column, 0)
        seen[column] = count + 1
        unique.append(column if count == 0 else f"{column}_{count + 1}")
    return unique


def _normalize_row(row: list[str], column_count: int) -> list[str]:
    normalized = row[:column_count]
    if len(normalized) < column_count:
        normalized.extend([""] * (column_count - len(normalized)))
    return normalized


def _extract_csv_rows(
    *,
    content: bytes,
    structure: StructureDetectionResult,
) -> list[list[str]] | None:
    if structure.structure_status != StructureDetectionStatus.DETECTED:
        return None
    delimiter = structure.delimiter.detected_delimiter if structure.delimiter else ","
    text = content[: 512 * 1024].decode("utf-8-sig", errors="replace")
    rows = [row for row in csv.reader(StringIO(text), delimiter=delimiter or ",")]
    start_index = structure.header_row_index or 0
    return [
        row
        for index, row in enumerate(rows[:MAX_CLEANING_SAMPLE_ROWS])
        if index >= start_index and (not structure.has_detected_header or index != start_index)
    ]


def _extract_excel_rows(
    *,
    content: bytes,
    extension: str,
    sheet_name: str,
    structure: StructureDetectionResult,
) -> list[list[str]] | None:
    if structure.structure_status != StructureDetectionStatus.DETECTED:
        return None
    try:
        rows = _sample_xlsx_rows(content, sheet_name) if extension == "xlsx" else _sample_xls_rows(content, sheet_name)
    except Exception:
        return None
    start_index = structure.header_row_index or 0
    return [
        row
        for index, row in enumerate(rows)
        if index >= start_index and (not structure.has_detected_header or index != start_index)
    ]


def _sample_xlsx_rows(content: bytes, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = []
    for row in worksheet.iter_rows(max_row=MAX_CLEANING_SAMPLE_ROWS, values_only=True):
        rows.append([_cell_to_text(cell) for cell in row])
    return rows


def _sample_xls_rows(content: bytes, sheet_name: str) -> list[list[str]]:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed")
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    sheet = workbook.sheet_by_name(sheet_name)
    rows: list[list[str]] = []
    for row_index in range(min(sheet.nrows, MAX_CLEANING_SAMPLE_ROWS)):
        rows.append([_cell_to_text(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)])
    return rows


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_selection_required_result(structure: StructureDetectionResult) -> CleaningPreviewResult:
    return CleaningPreviewResult(
        cleaning_status=CleaningStatus.SHEET_SELECTION_REQUIRED,
        original_filename=structure.original_filename,
        safe_filename=structure.safe_filename,
        detected_extension=structure.detected_extension,
        before_summary=CleaningSummaryBefore(row_count=0, column_count=0),
        after_summary=CleaningSummaryAfter(row_count=0, column_count=0),
        warnings=[
            CleaningPreviewWarning(
                code="sheet_selection_required",
                message="Select an Excel sheet before generating a cleaning preview.",
                severity=CleaningWarningSeverity.WARNING,
            )
        ],
        next_step=CleaningNextStep.SELECT_EXCEL_SHEET,
    )


def _rejected_result_from_structure(structure: StructureDetectionResult) -> CleaningPreviewResult:
    message = (
        "File is empty. Upload a non-empty tabular file."
        if structure.file_size_bytes == 0
        else (structure.warnings[0].message if structure.warnings else "File cannot be cleaned in this milestone.")
    )
    return CleaningPreviewResult(
        cleaning_status=CleaningStatus.REJECTED,
        original_filename=structure.original_filename,
        safe_filename=structure.safe_filename,
        detected_extension=structure.detected_extension,
        selected_sheet_name=structure.selected_sheet_name,
        before_summary=CleaningSummaryBefore(row_count=0, column_count=0),
        after_summary=CleaningSummaryAfter(row_count=0, column_count=0),
        warnings=[
            CleaningPreviewWarning(
                code=structure.warnings[0].code if structure.warnings else "cleaning_preview_rejected",
                message=message,
                severity=CleaningWarningSeverity.ERROR,
            )
        ],
        next_step=CleaningNextStep.UPLOAD_SUPPORTED_FILE,
    )
