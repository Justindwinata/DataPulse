import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None

from datapulse_api.models import CleaningRuleCode, StructureDetectionStatus
from datapulse_api.services.cleaning_engine import (
    _apply_rule,
    _cell_to_text,
    _normalize_row,
)
from datapulse_api.services.csv_structure_detection import (
    CSV_LIKE_EXTENSIONS,
    EXCEL_EXTENSIONS,
    detect_csv_like_structure,
)
from datapulse_api.services.excel_structure_detection import (
    detect_excel_sheet_preview,
    detect_excel_workbook,
)
from datapulse_api.services.file_validation import detect_extension, sanitize_filename


@dataclass(frozen=True)
class CleanedCsvExport:
    filename: str
    content: bytes
    content_type: str = "text/csv; charset=utf-8"


class CleanedCsvExportError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def export_cleaned_csv(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    rules: Iterable[CleaningRuleCode],
    sheet_name: str | None = None,
) -> CleanedCsvExport:
    safe_filename = sanitize_filename(filename)
    extension = detect_extension(safe_filename) or "unknown"
    selected_rules = list(rules)

    if extension in CSV_LIKE_EXTENSIONS:
        structure = detect_csv_like_structure(
            filename=filename,
            content_type=content_type,
            content=content,
        )
        if structure.structure_status != StructureDetectionStatus.DETECTED or structure.preview is None:
            _raise_from_structure(structure)
        rows = _extract_all_csv_rows(content=content, structure=structure)
        return _export_from_table(
            source_filename=safe_filename,
            columns=list(structure.preview.columns),
            rows=rows,
            rules=selected_rules,
        )

    if extension in EXCEL_EXTENSIONS:
        if not sheet_name:
            workbook = detect_excel_workbook(
                filename=filename,
                content_type=content_type,
                content=content,
            )
            if workbook.structure_status == StructureDetectionStatus.SHEET_SELECTION_REQUIRED:
                raise CleanedCsvExportError(
                    "sheet_selection_required",
                    "Select an Excel sheet before exporting cleaned CSV.",
                )
            _raise_from_structure(workbook)
        structure = detect_excel_sheet_preview(
            filename=filename,
            content_type=content_type,
            content=content,
            sheet_name=sheet_name,
        )
        if structure.structure_status != StructureDetectionStatus.DETECTED or structure.preview is None:
            _raise_from_structure(structure)
        rows = _extract_all_excel_rows(
            content=content,
            extension=extension,
            sheet_name=sheet_name,
            header_row_index=structure.header_row_index,
            has_detected_header=structure.has_detected_header,
        )
        return _export_from_table(
            source_filename=safe_filename,
            columns=list(structure.preview.columns),
            rows=rows,
            rules=selected_rules,
        )

    structure = detect_csv_like_structure(
        filename=filename,
        content_type=content_type,
        content=content,
    )
    _raise_from_structure(structure)


def _export_from_table(
    *,
    source_filename: str,
    columns: list[str],
    rows: list[list[str]],
    rules: list[CleaningRuleCode],
) -> CleanedCsvExport:
    normalized_rows = [_normalize_row(row, len(columns)) for row in rows]
    current_columns = list(columns)
    current_rows = [list(row) for row in normalized_rows]

    for rule in rules:
        current_columns, current_rows, _ = _apply_rule(rule, current_columns, current_rows)

    csv_text = _render_csv(current_columns, current_rows)
    return CleanedCsvExport(
        filename=_cleaned_filename(source_filename),
        content=csv_text.encode("utf-8"),
    )


def _extract_all_csv_rows(*, content: bytes, structure) -> list[list[str]]:
    delimiter = structure.delimiter.detected_delimiter if structure.delimiter else ","
    text = content.decode("utf-8-sig", errors="replace")
    parsed_rows = [row for row in csv.reader(StringIO(text), delimiter=delimiter or ",")]
    start_index = structure.header_row_index or 0
    return [
        row
        for index, row in enumerate(parsed_rows)
        if index >= start_index and (not structure.has_detected_header or index != start_index)
    ]


def _extract_all_excel_rows(
    *,
    content: bytes,
    extension: str,
    sheet_name: str,
    header_row_index: int | None,
    has_detected_header: bool,
) -> list[list[str]]:
    rows = (
        _extract_all_xlsx_rows(content, sheet_name)
        if extension == "xlsx"
        else _extract_all_xls_rows(content, sheet_name)
    )
    start_index = header_row_index or 0
    return [
        row
        for index, row in enumerate(rows)
        if index >= start_index and (not has_detected_header or index != start_index)
    ]


def _extract_all_xlsx_rows(content: bytes, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    return [[_cell_to_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]


def _extract_all_xls_rows(content: bytes, sheet_name: str) -> list[list[str]]:
    if xlrd is None:
        raise CleanedCsvExportError("xls_not_available", "XLS support is not available.")
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    sheet = workbook.sheet_by_name(sheet_name)
    return [
        [_cell_to_text(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
        for row_index in range(sheet.nrows)
    ]


def _render_csv(columns: list[str], rows: list[list[str]]) -> str:
    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    writer.writerow(columns)
    writer.writerows(rows)
    return output.getvalue()


def _cleaned_filename(filename: str) -> str:
    stem = Path(filename).stem or "cleaned_data"
    safe_stem = "".join(character if character.isalnum() or character in {"-", "_"} else "_" for character in stem)
    safe_stem = safe_stem.strip("._-") or "cleaned_data"
    return f"{safe_stem}_cleaned.csv"


def _raise_from_structure(structure) -> None:
    if structure.file_size_bytes == 0:
        raise CleanedCsvExportError("empty_file", "File is empty. Upload a non-empty tabular file.")
    warning = structure.warnings[0] if structure.warnings else None
    raise CleanedCsvExportError(
        warning.code if warning else "export_rejected",
        warning.message if warning else "File cannot be exported as cleaned CSV.",
    )
