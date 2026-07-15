import csv
import re
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency is declared, fallback is defensive.
    xlrd = None

from datapulse_api.models import (
    ColumnProfile,
    DataQualityIssue,
    DataQualityResult,
    InferredColumnType,
    IssueSeverity,
    QualityNextStep,
    QualityStatus,
    SeverityCounts,
    StructureDetectionResult,
    StructureDetectionStatus,
)
from datapulse_api.services.csv_structure_detection import (
    CSV_LIKE_EXTENSIONS,
    EXCEL_EXTENSIONS,
    detect_csv_like_structure,
)
from datapulse_api.services.excel_structure_detection import (
    detect_excel_sheet_preview,
    detect_excel_workbook,
)
from datapulse_api.services.file_validation import detect_extension, sanitize_filename

HIGH_MISSING_THRESHOLD = 50.0
MAX_SAMPLE_VALUES = 5
MAX_QUALITY_SAMPLE_ROWS = 50
MISSING_TOKEN_VALUES = {"unknown", "error", "n/a", "na", "null", "none", "nan", "-"}
DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%m-%d-%Y",
    "%Y-%m-%d %H:%M:%S",
)
GENERIC_COLUMN_RE = re.compile(r"^column_\d+$")


def detect_data_quality(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    sheet_name: str | None = None,
) -> DataQualityResult:
    safe_filename = sanitize_filename(filename)
    extension = detect_extension(safe_filename) or "unknown"

    if extension in CSV_LIKE_EXTENSIONS:
        structure = detect_csv_like_structure(
            filename=filename,
            content_type=content_type,
            content=content,
        )
        rows = _extract_csv_quality_rows(content=content, structure=structure)
        return _quality_from_structure(structure, rows=rows)

    if extension in EXCEL_EXTENSIONS:
        if not sheet_name:
            workbook_result = detect_excel_workbook(
                filename=filename,
                content_type=content_type,
                content=content,
            )
            if workbook_result.structure_status == StructureDetectionStatus.SHEET_SELECTION_REQUIRED:
                return _sheet_selection_required_result(workbook_result)
            return _rejected_result_from_structure(workbook_result)

        structure = detect_excel_sheet_preview(
            filename=filename,
            content_type=content_type,
            content=content,
            sheet_name=sheet_name,
        )
        rows = _extract_excel_quality_rows(
            content=content,
            extension=extension,
            sheet_name=sheet_name,
            structure=structure,
        )
        return _quality_from_structure(structure, rows=rows)

    structure = detect_csv_like_structure(
        filename=filename,
        content_type=content_type,
        content=content,
    )
    return _rejected_result_from_structure(structure)


def _quality_from_structure(
    structure: StructureDetectionResult,
    rows: list[list[str]] | None = None,
) -> DataQualityResult:
    if structure.structure_status != StructureDetectionStatus.DETECTED or structure.preview is None:
        return _rejected_result_from_structure(structure)

    columns = structure.preview.columns
    source_rows = rows if rows is not None else structure.preview.rows
    normalized_rows = [_normalize_row(row, len(columns)) for row in source_rows]
    profiles = _build_column_profiles(columns, normalized_rows)
    issues = _detect_issues(
        columns=columns,
        rows=normalized_rows,
        profiles=profiles,
        structure=structure,
    )
    profiles = _apply_column_issue_codes(profiles, issues)
    severity_counts = _build_severity_counts(issues)

    return DataQualityResult(
        quality_status=QualityStatus.PROFILED,
        original_filename=structure.original_filename,
        safe_filename=structure.safe_filename,
        detected_extension=structure.detected_extension,
        selected_sheet_name=structure.selected_sheet_name,
        sampled_row_count=len(normalized_rows),
        detected_column_count=len(columns),
        total_issue_count=len(issues),
        severity_counts=severity_counts,
        quality_score=_calculate_quality_score(severity_counts),
        issues=issues,
        columns=profiles,
        next_step=QualityNextStep.CLEANING_RULES,
    )


def _build_column_profiles(columns: list[str], rows: list[list[str]]) -> list[ColumnProfile]:
    profiles: list[ColumnProfile] = []
    row_count = len(rows)
    for column_index, column_name in enumerate(columns):
        values = [row[column_index] if column_index < len(row) else "" for row in rows]
        stripped_values = [value.strip() for value in values]
        non_empty_values = [value for value in stripped_values if value and not _is_missing_token(value)]
        missing_count = row_count - len(non_empty_values)
        unique_values = list(dict.fromkeys(non_empty_values))
        profiles.append(
            ColumnProfile(
                column_name=column_name,
                column_index=column_index,
                non_empty_count=len(non_empty_values),
                missing_count=missing_count,
                missing_percentage=round((missing_count / row_count) * 100, 2)
                if row_count
                else 0,
                inferred_type=_infer_column_type(non_empty_values),
                unique_count=len(set(non_empty_values)),
                sample_values=unique_values[:MAX_SAMPLE_VALUES],
                issues=[],
            )
        )
    return profiles


def _detect_issues(
    *,
    columns: list[str],
    rows: list[list[str]],
    profiles: list[ColumnProfile],
    structure: StructureDetectionResult,
) -> list[DataQualityIssue]:
    issues: list[DataQualityIssue] = []

    missing_columns = [profile.column_name for profile in profiles if profile.missing_count > 0]
    missing_row_count = sum(1 for row in rows if any(not cell.strip() for cell in row))
    if missing_columns:
        issues.append(
            DataQualityIssue(
                code="missing_values",
                title="Missing values detected",
                message="Some columns contain empty values in the sampled data.",
                severity=IssueSeverity.WARNING,
                affected_columns=missing_columns,
                affected_row_count=missing_row_count,
            )
        )

    empty_row_count = sum(1 for row in rows if not any(cell.strip() for cell in row))
    if not empty_row_count and any(
        warning.code == "empty_rows_in_sample" for warning in structure.warnings
    ):
        empty_row_count = 1
    if empty_row_count:
        issues.append(
            DataQualityIssue(
                code="empty_rows",
                title="Empty rows detected",
                message="Rows with no values were found in the sampled data.",
                severity=IssueSeverity.WARNING,
                affected_row_count=empty_row_count,
                suggested_cleaning_rule="remove_empty_rows",
            )
        )

    empty_columns = [
        profile.column_name
        for profile in profiles
        if profile.inferred_type == InferredColumnType.EMPTY
    ]
    if empty_columns:
        issues.append(
            DataQualityIssue(
                code="empty_columns",
                title="Empty columns detected",
                message="One or more columns have no values in the sampled data.",
                severity=IssueSeverity.CRITICAL,
                affected_columns=empty_columns,
                suggested_cleaning_rule="drop_empty_columns",
            )
        )

    duplicate_row_count = _count_duplicate_rows(rows)
    if duplicate_row_count:
        issues.append(
            DataQualityIssue(
                code="duplicate_rows",
                title="Duplicate rows detected",
                message="Repeated rows were found in the sampled data.",
                severity=IssueSeverity.WARNING,
                affected_row_count=duplicate_row_count,
                suggested_cleaning_rule="remove_duplicate_rows",
            )
        )

    duplicate_columns = _duplicate_column_names(columns)
    if duplicate_columns:
        issues.append(
            DataQualityIssue(
                code="duplicate_column_names",
                title="Duplicate column names detected",
                message="The sampled header contains repeated column names.",
                severity=IssueSeverity.CRITICAL,
                affected_columns=duplicate_columns,
                suggested_cleaning_rule="standardize_column_names",
            )
        )

    generated_columns = [column for column in columns if GENERIC_COLUMN_RE.match(column)]
    if structure.generated_column_names or generated_columns:
        issues.append(
            DataQualityIssue(
                code="missing_column_names",
                title="Missing column names detected",
                message="Some column names were generated because headers were missing or incomplete.",
                severity=IssueSeverity.WARNING,
                affected_columns=generated_columns,
                suggested_cleaning_rule="generate_missing_column_names",
            )
        )

    messy_columns = [column for column in columns if _is_messy_column_name(column)]
    if messy_columns:
        issues.append(
            DataQualityIssue(
                code="messy_column_names",
                title="Messy column names detected",
                message="Some column names contain spacing or characters that may be awkward for analysis.",
                severity=IssueSeverity.INFO,
                affected_columns=messy_columns,
                suggested_cleaning_rule="standardize_column_names",
            )
        )

    if any(warning.code == "inconsistent_row_widths" for warning in structure.warnings):
        issues.append(
            DataQualityIssue(
                code="inconsistent_row_widths",
                title="Inconsistent row widths detected",
                message="Rows with different column counts were detected before preview normalization.",
                severity=IssueSeverity.CRITICAL,
            )
        )

    whitespace_columns, whitespace_rows = _detect_whitespace(rows, columns)
    if whitespace_columns:
        issues.append(
            DataQualityIssue(
                code="leading_trailing_whitespace",
                title="Leading or trailing whitespace detected",
                message="Some sampled values contain whitespace around the visible value.",
                severity=IssueSeverity.INFO,
                affected_columns=whitespace_columns,
                affected_row_count=whitespace_rows,
                suggested_cleaning_rule="trim_whitespace",
            )
        )

    placeholder_columns, placeholder_rows = _detect_placeholder_missing_tokens(rows, columns)
    if placeholder_columns:
        issues.append(
            DataQualityIssue(
                code="placeholder_missing_values",
                title="Placeholder missing values detected",
                message="Some cells contain placeholders such as UNKNOWN, ERROR, N/A, NULL, or dash values.",
                severity=IssueSeverity.WARNING,
                affected_columns=placeholder_columns,
                affected_row_count=placeholder_rows,
                suggested_cleaning_rule="normalize_missing_tokens",
            )
        )

    invalid_numeric_columns, invalid_numeric_rows = _detect_invalid_numeric_values(rows, columns)
    if invalid_numeric_columns:
        issues.append(
            DataQualityIssue(
                code="invalid_numeric_values",
                title="Invalid numeric values detected",
                message="Numeric-like columns contain non-numeric placeholders or invalid values.",
                severity=IssueSeverity.WARNING,
                affected_columns=invalid_numeric_columns,
                affected_row_count=invalid_numeric_rows,
                suggested_cleaning_rule="clean_numeric_values",
            )
        )

    invalid_date_columns, invalid_date_rows = _detect_invalid_date_values(rows, columns)
    if invalid_date_columns:
        issues.append(
            DataQualityIssue(
                code="invalid_date_values",
                title="Invalid date values detected",
                message="Date-like columns contain unparseable values or placeholders.",
                severity=IssueSeverity.WARNING,
                affected_columns=invalid_date_columns,
                affected_row_count=invalid_date_rows,
                suggested_cleaning_rule="clean_date_values",
            )
        )

    category_columns = _detect_category_text_columns(rows, columns)
    if category_columns:
        issues.append(
            DataQualityIssue(
                code="category_text_inconsistency",
                title="Category text can be standardized",
                message="Some text columns look like repeated categories that can be standardized safely.",
                severity=IssueSeverity.INFO,
                affected_columns=category_columns,
                suggested_cleaning_rule="standardize_category_text",
            )
        )

    inconsistent_total_rows = _detect_inconsistent_line_totals(rows, columns)
    if inconsistent_total_rows:
        issues.append(
            DataQualityIssue(
                code="inconsistent_line_totals",
                title="Line totals can be recalculated",
                message="Quantity, price, and total columns were detected; line totals can be recalculated where quantity and price are valid.",
                severity=IssueSeverity.INFO,
                affected_columns=_line_total_columns(columns),
                affected_row_count=inconsistent_total_rows,
                suggested_cleaning_rule="recalculate_line_totals",
            )
        )

    mixed_columns = [
        profile.column_name
        for profile in profiles
        if profile.inferred_type == InferredColumnType.MIXED
    ]
    if mixed_columns:
        issues.append(
            DataQualityIssue(
                code="mixed_type_values",
                title="Mixed value types detected",
                message="Some columns contain values that look like different data types.",
                severity=IssueSeverity.WARNING,
                affected_columns=mixed_columns,
            )
        )

    numeric_columns = [
        profile.column_name
        for profile in profiles
        if profile.inferred_type == InferredColumnType.NUMBER
    ]
    if numeric_columns:
        issues.append(
            DataQualityIssue(
                code="numeric_values_as_text",
                title="Possible numeric columns stored as text",
                message="Some sampled columns contain numeric-looking text values.",
                severity=IssueSeverity.INFO,
                affected_columns=numeric_columns,
                suggested_cleaning_rule="clean_numeric_values",
            )
        )

    date_columns = [
        profile.column_name
        for profile in profiles
        if profile.inferred_type == InferredColumnType.DATE
    ]
    if date_columns:
        issues.append(
            DataQualityIssue(
                code="date_values_as_text",
                title="Possible date columns stored as text",
                message="Some sampled columns contain date-looking text values.",
                severity=IssueSeverity.INFO,
                affected_columns=date_columns,
                suggested_cleaning_rule="clean_date_values",
            )
        )

    high_missing_columns = [
        profile.column_name
        for profile in profiles
        if profile.missing_count > 0 and profile.missing_percentage >= HIGH_MISSING_THRESHOLD
    ]
    if high_missing_columns:
        issues.append(
            DataQualityIssue(
                code="high_missing_column",
                title="High missing-value columns detected",
                message="Some columns are missing values in at least half of the sampled rows.",
                severity=IssueSeverity.WARNING,
                affected_columns=high_missing_columns,
            )
        )

    return _deduplicate_issues(issues)


def _apply_column_issue_codes(
    profiles: list[ColumnProfile],
    issues: list[DataQualityIssue],
) -> list[ColumnProfile]:
    updated: list[ColumnProfile] = []
    for profile in profiles:
        issue_codes = [
            issue.code
            for issue in issues
            if profile.column_name in issue.affected_columns
        ]
        updated.append(profile.model_copy(update={"issues": issue_codes}))
    return updated


def _infer_column_type(values: list[str]) -> InferredColumnType:
    if not values:
        return InferredColumnType.EMPTY

    detected_types = [_detect_value_type(value) for value in values]
    type_counts = Counter(detected_types)
    if len(type_counts) == 1:
        return detected_types[0]
    dominant_type, dominant_count = type_counts.most_common(1)[0]
    if dominant_count / len(values) >= 0.8 and dominant_type != InferredColumnType.TEXT:
        return dominant_type
    return InferredColumnType.MIXED


def _detect_value_type(value: str) -> InferredColumnType:
    normalized = value.strip()
    if _is_missing_token(normalized):
        return InferredColumnType.EMPTY
    if _is_boolean(normalized):
        return InferredColumnType.BOOLEAN
    if _is_number(normalized):
        return InferredColumnType.NUMBER
    if _is_date(normalized):
        return InferredColumnType.DATE
    return InferredColumnType.TEXT


def _is_number(value: str) -> bool:
    normalized = value.strip().replace(",", "")
    if normalized.startswith("$"):
        normalized = normalized[1:]
    if not normalized:
        return False
    try:
        float(normalized)
    except ValueError:
        return False
    return True


def _is_missing_token(value: str) -> bool:
    normalized = value.strip().lower()
    return not normalized or normalized in MISSING_TOKEN_VALUES


def _is_date(value: str) -> bool:
    normalized = value.strip()
    for date_format in DATE_FORMATS:
        try:
            datetime.strptime(normalized, date_format)
        except ValueError:
            continue
        return True
    return False


def _parse_number(value: str) -> Decimal | None:
    normalized = value.strip().replace(",", "")
    if normalized.startswith("$"):
        normalized = normalized[1:]
    if not normalized:
        return None
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def _normalize_column_name(column: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", column.strip().lower()).strip("_")


def _is_numeric_candidate(column: str, values: list[str]) -> bool:
    normalized_column = _normalize_column_name(column)
    if any(token in normalized_column for token in ("quantity", "amount", "price", "total", "spent", "cost")):
        return True
    non_missing = [value for value in values if not _is_missing_token(value)]
    if not non_missing:
        return False
    numeric_count = sum(1 for value in non_missing if _parse_number(value) is not None)
    return numeric_count / len(non_missing) >= 0.8


def _is_date_candidate(column: str, values: list[str]) -> bool:
    normalized_column = _normalize_column_name(column)
    if "date" in normalized_column or normalized_column.endswith("_at"):
        return True
    non_missing = [value for value in values if not _is_missing_token(value)]
    if not non_missing:
        return False
    date_count = sum(1 for value in non_missing if _is_date(value))
    return date_count / len(non_missing) >= 0.8


def _detect_placeholder_missing_tokens(rows: list[list[str]], columns: list[str]) -> tuple[list[str], int]:
    affected_indexes: set[int] = set()
    affected_rows = 0
    for row in rows:
        row_has_placeholder = False
        for column_index, cell in enumerate(row):
            normalized = cell.strip().lower()
            if normalized in MISSING_TOKEN_VALUES:
                affected_indexes.add(column_index)
                row_has_placeholder = True
        if row_has_placeholder:
            affected_rows += 1
    return [columns[index] for index in sorted(affected_indexes)], affected_rows


def _detect_invalid_numeric_values(rows: list[list[str]], columns: list[str]) -> tuple[list[str], int]:
    affected_indexes: set[int] = set()
    affected_rows: set[int] = set()
    for column_index, column in enumerate(columns):
        values = [row[column_index] if column_index < len(row) else "" for row in rows]
        if not _is_numeric_candidate(column, values):
            continue
        for row_index, value in enumerate(values):
            if _is_missing_token(value):
                if value.strip():
                    affected_indexes.add(column_index)
                    affected_rows.add(row_index)
                continue
            if _parse_number(value) is None:
                affected_indexes.add(column_index)
                affected_rows.add(row_index)
    return [columns[index] for index in sorted(affected_indexes)], len(affected_rows)


def _detect_invalid_date_values(rows: list[list[str]], columns: list[str]) -> tuple[list[str], int]:
    affected_indexes: set[int] = set()
    affected_rows: set[int] = set()
    for column_index, column in enumerate(columns):
        values = [row[column_index] if column_index < len(row) else "" for row in rows]
        if not _is_date_candidate(column, values):
            continue
        for row_index, value in enumerate(values):
            if _is_missing_token(value):
                if value.strip():
                    affected_indexes.add(column_index)
                    affected_rows.add(row_index)
                continue
            if not _is_date(value):
                affected_indexes.add(column_index)
                affected_rows.add(row_index)
    return [columns[index] for index in sorted(affected_indexes)], len(affected_rows)


def _detect_category_text_columns(rows: list[list[str]], columns: list[str]) -> list[str]:
    category_columns: list[str] = []
    for column_index, column in enumerate(columns):
        normalized_column = _normalize_column_name(column)
        if "id" in normalized_column or _is_numeric_candidate(column, [row[column_index] for row in rows]):
            continue
        values = [row[column_index] if column_index < len(row) else "" for row in rows]
        if _is_date_candidate(column, values):
            continue
        non_missing = [value.strip() for value in values if not _is_missing_token(value)]
        if not non_missing:
            continue
        unique_normalized = {re.sub(r"\s+", " ", value).casefold() for value in non_missing}
        has_cleanable_values = any(_clean_category_preview(value) != value for value in values)
        if has_cleanable_values and len(unique_normalized) <= max(20, int(len(non_missing) * 0.4)):
            category_columns.append(column)
    return category_columns


def _clean_category_preview(value: str) -> str:
    if _is_missing_token(value):
        return ""
    collapsed = re.sub(r"\s+", " ", value.strip())
    return " ".join("-".join(part.capitalize() for part in word.split("-")) for word in collapsed.split(" "))


def _line_total_column_indexes(columns: list[str]) -> tuple[int, int, int] | None:
    normalized = [_normalize_column_name(column) for column in columns]
    quantity = next((index for index, column in enumerate(normalized) if column in {"quantity", "qty"}), None)
    price = next((index for index, column in enumerate(normalized) if "price" in column and "unit" in column), None)
    total = next((index for index, column in enumerate(normalized) if "total" in column and ("spent" in column or "amount" in column or column == "total")), None)
    if quantity is None or price is None or total is None:
        return None
    return quantity, price, total


def _line_total_columns(columns: list[str]) -> list[str]:
    indexes = _line_total_column_indexes(columns)
    return [columns[index] for index in indexes] if indexes else []


def _detect_inconsistent_line_totals(rows: list[list[str]], columns: list[str]) -> int:
    indexes = _line_total_column_indexes(columns)
    if indexes is None:
        return 0
    quantity_index, price_index, total_index = indexes
    affected_rows = 0
    recognized_rows = 0
    for row in rows:
        quantity = _parse_number(row[quantity_index]) if quantity_index < len(row) else None
        price = _parse_number(row[price_index]) if price_index < len(row) else None
        total = _parse_number(row[total_index]) if total_index < len(row) else None
        if quantity is None or price is None:
            continue
        recognized_rows += 1
        expected = quantity * price
        if total is None or total != expected:
            affected_rows += 1
    return affected_rows if recognized_rows else 0


def _is_boolean(value: str) -> bool:
    return value.strip().lower() in {"true", "false", "yes", "no"}


def _is_messy_column_name(column: str) -> bool:
    return column != column.strip() or bool(re.search(r"[^a-z0-9_]", column))


def _duplicate_column_names(columns: list[str]) -> list[str]:
    counts = Counter(columns)
    return [column for column, count in counts.items() if count > 1]


def _count_duplicate_rows(rows: list[list[str]]) -> int:
    normalized = [
        tuple(cell.strip() for cell in row)
        for row in rows
        if any(cell.strip() for cell in row)
    ]
    counts = Counter(normalized)
    return sum(count - 1 for count in counts.values() if count > 1)


def _detect_whitespace(rows: list[list[str]], columns: list[str]) -> tuple[list[str], int]:
    affected_indexes: set[int] = set()
    affected_rows = 0
    for row in rows:
        row_has_whitespace = False
        for column_index, cell in enumerate(row):
            if cell and cell != cell.strip():
                affected_indexes.add(column_index)
                row_has_whitespace = True
        if row_has_whitespace:
            affected_rows += 1
    return [columns[index] for index in sorted(affected_indexes)], affected_rows


def _build_severity_counts(issues: list[DataQualityIssue]) -> SeverityCounts:
    counts = Counter(issue.severity for issue in issues)
    return SeverityCounts(
        info=counts[IssueSeverity.INFO],
        warning=counts[IssueSeverity.WARNING],
        critical=counts[IssueSeverity.CRITICAL],
    )


def _calculate_quality_score(severity_counts: SeverityCounts) -> int:
    score = (
        100
        - severity_counts.critical * 15
        - severity_counts.warning * 7
        - severity_counts.info * 2
    )
    return max(0, min(100, score))


def _normalize_row(row: list[str], column_count: int) -> list[str]:
    normalized = row[:column_count]
    if len(normalized) < column_count:
        normalized.extend([""] * (column_count - len(normalized)))
    return normalized


def _deduplicate_issues(issues: list[DataQualityIssue]) -> list[DataQualityIssue]:
    seen: set[str] = set()
    unique: list[DataQualityIssue] = []
    for issue in issues:
        if issue.code in seen:
            continue
        seen.add(issue.code)
        unique.append(issue)
    return unique


def _sheet_selection_required_result(structure: StructureDetectionResult) -> DataQualityResult:
    message = "Select an Excel sheet before running data quality profiling."
    return DataQualityResult(
        quality_status=QualityStatus.SHEET_SELECTION_REQUIRED,
        original_filename=structure.original_filename,
        safe_filename=structure.safe_filename,
        detected_extension=structure.detected_extension,
        sampled_row_count=0,
        detected_column_count=0,
        quality_score=100,
        messages=[message],
        next_step=QualityNextStep.SELECT_EXCEL_SHEET,
    )


def _rejected_result_from_structure(structure: StructureDetectionResult) -> DataQualityResult:
    messages = [warning.message for warning in structure.warnings]
    if structure.file_size_bytes == 0 or any(
        warning.code == "empty_file" for warning in structure.warnings
    ):
        messages = ["File is empty. Upload a non-empty tabular file."]
    return DataQualityResult(
        quality_status=QualityStatus.REJECTED,
        original_filename=structure.original_filename,
        safe_filename=structure.safe_filename,
        detected_extension=structure.detected_extension,
        selected_sheet_name=structure.selected_sheet_name,
        sampled_row_count=0,
        detected_column_count=0,
        quality_score=100,
        messages=messages,
        next_step=QualityNextStep.UPLOAD_SUPPORTED_FILE,
    )


def _extract_csv_quality_rows(
    *,
    content: bytes,
    structure: StructureDetectionResult,
) -> list[list[str]] | None:
    if structure.structure_status != StructureDetectionStatus.DETECTED:
        return None
    delimiter = structure.delimiter.detected_delimiter if structure.delimiter else ","
    if delimiter is None:
        delimiter = ","
    text = content[: 512 * 1024].decode("utf-8-sig", errors="replace")
    rows = [row for row in csv.reader(StringIO(text), delimiter=delimiter)]
    if not rows:
        return []
    start_index = structure.header_row_index or 0
    data_rows = [
        row
        for index, row in enumerate(rows[:MAX_QUALITY_SAMPLE_ROWS])
        if index >= start_index and (not structure.has_detected_header or index != start_index)
    ]
    return data_rows


def _extract_excel_quality_rows(
    *,
    content: bytes,
    extension: str,
    sheet_name: str,
    structure: StructureDetectionResult,
) -> list[list[str]] | None:
    if structure.structure_status != StructureDetectionStatus.DETECTED:
        return None
    try:
        rows = (
            _sample_xlsx_quality_rows(content, sheet_name)
            if extension == "xlsx"
            else _sample_xls_quality_rows(content, sheet_name)
        )
    except Exception:
        return None
    start_index = structure.header_row_index or 0
    return [
        row
        for index, row in enumerate(rows)
        if index >= start_index and (not structure.has_detected_header or index != start_index)
    ]


def _sample_xlsx_quality_rows(content: bytes, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(max_row=MAX_QUALITY_SAMPLE_ROWS, values_only=True):
        rows.append([_quality_cell_to_text(cell) for cell in row])
    return rows


def _sample_xls_quality_rows(content: bytes, sheet_name: str) -> list[list[str]]:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed")
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    sheet = workbook.sheet_by_name(sheet_name)
    rows: list[list[str]] = []
    for row_index in range(min(sheet.nrows, MAX_QUALITY_SAMPLE_ROWS)):
        rows.append(
            [
                _quality_cell_to_text(sheet.cell_value(row_index, column_index))
                for column_index in range(sheet.ncols)
            ]
        )
    return rows


def _quality_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)
