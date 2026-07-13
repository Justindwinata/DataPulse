from io import BytesIO

from openpyxl import Workbook

from datapulse_api.models import StructureDetectionStatus
from datapulse_api.services.excel_structure_detection import detect_excel_workbook


def make_xlsx_workbook(sheet_names: list[str]) -> bytes:
    workbook = Workbook()
    workbook.active.title = sheet_names[0]
    for sheet_name in sheet_names[1:]:
        workbook.create_sheet(sheet_name)
    workbook[sheet_names[0]].append(["customer", "amount"])
    workbook[sheet_names[0]].append(["Ari", 1200])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_xlsx_sheet_discovery_returns_workbook_metadata() -> None:
    content = make_xlsx_workbook(["Sales"])

    result = detect_excel_workbook(
        filename="sales.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=content,
    )

    assert result.structure_status == StructureDetectionStatus.SHEET_SELECTION_REQUIRED
    assert result.workbook is not None
    assert result.workbook.workbook_type == "xlsx"
    assert result.workbook.sheet_names == ["Sales"]
    assert result.workbook.sheet_count == 1
    assert result.workbook.default_sheet_name == "Sales"
    assert result.workbook.sheets[0].max_row == 2
    assert result.workbook.sheets[0].max_column == 2


def test_xlsx_multiple_sheets_are_discovered_in_order() -> None:
    content = make_xlsx_workbook(["Sales", "Customers", "Raw Export"])

    result = detect_excel_workbook(
        filename="multi.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=content,
    )

    assert result.workbook is not None
    assert result.workbook.sheet_names == ["Sales", "Customers", "Raw Export"]
    assert result.workbook.sheet_count == 3
    assert result.workbook.default_sheet_name == "Sales"


def test_xlsx_sheet_discovery_returns_excel_limitation_warning() -> None:
    content = make_xlsx_workbook(["Sales"])

    result = detect_excel_workbook(
        filename="sales.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=content,
    )

    assert result.warnings[0].code == "excel_formatting_not_preserved"
    assert "formatting" in result.warnings[0].message


def test_empty_excel_file_is_rejected() -> None:
    result = detect_excel_workbook(
        filename="empty.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=b"",
    )

    assert result.structure_status == StructureDetectionStatus.REJECTED
    assert result.warnings[0].code == "empty_file"


def test_non_excel_file_is_rejected_by_excel_discovery_service() -> None:
    result = detect_excel_workbook(
        filename="document.pdf",
        content_type="application/pdf",
        content=b"%PDF",
    )

    assert result.structure_status == StructureDetectionStatus.REJECTED
    assert result.warnings[0].code == "file_not_excel"
